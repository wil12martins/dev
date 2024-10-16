import streamlit as st
import pandas as pd
import os
from zipfile import ZipFile
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# Função para processar cada CSV carregado
def processar_csv(df):
    # Colunas desejadas
    colunas_desejadas = [
        "QUADRA", "FACE", "LOGRADOURO", "ENDERECO", "PONTO DE REFERENCIA",
        "LOCALIDADE", "ESPECIE", "NOME_RESPONSAVEL_1", "TELEFONE_1"
    ]
    


    # Garante que SEQ_UV, FACE e QUADRA são tratadas como números inteiros
    df['SEQ_UV'] = df['SEQ_UV'].astype(int)
    df['FACE'] = df['FACE'].astype(int)
    df['QUADRA'] = df['QUADRA'].astype(int)

    # Ordena pelos campos SEQ_UV, FACE e QUADRA
    df = df.sort_values(by=['QUADRA', 'FACE', 'SEQ_UV'], ascending=[True, True, True])
    # Filtra apenas as colunas desejadas
    df = df[colunas_desejadas]
    # Função para transformar a coluna ESPECIE
    def transformar_especie(especie):
        if pd.isna(especie) or especie.strip() == "":
            return especie
        return ''.join(word[0].upper() for word in especie.split())

    # Aplica a transformação à coluna ESPECIE
    df['ESPECIE'] = df['ESPECIE'].apply(transformar_especie)
    df.rename(columns={'QUADRA': 'Q', 'FACE': 'F'}, inplace=True)
    return df

# Função para converter CSV em Excel
def csv_para_excel(df, filename):
    # Cria um novo Workbook
    wb = Workbook()
    ws = wb.active

    # Define o nome da aba com base no nome do arquivo
    sheet_name = filename.split('_')[-1].replace('.csv', '')
    ws.title = sheet_name

    # Adiciona os cabeçalhos do DataFrame ao Excel
    for c_idx, column in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=column)
        ws.cell(row=1, column=c_idx).font = Font(size=6)  # Tamanho da fonte da primeira linha
    
    # Define a altura da linha do cabeçalho
    ws.row_dimensions[1].height = 20  # Altura da linha do cabeçalho

    # Adiciona os dados do DataFrame ao Excel
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.row_dimensions[r_idx].height = 30  # Altura das linhas de dados

    # Aplica formatação (bordas, fonte, etc.)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    for cell in ws[1]:
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = bottom_border

    # Ajusta as margens e orientação
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5 / 2.54
    ws.page_margins.right = 0.5 / 2.54
    ws.page_margins.top = 1.0 / 2.54
    ws.page_margins.bottom = 1.0 / 2.54

    # Congela a primeira linha
    ws.freeze_panes = 'A2'

    # Repetir a primeira linha em todas as páginas ao imprimir
    ws.print_title_rows = '1:1'  # Define a primeira linha para ser repetida em todas as páginas

    # Adiciona o nome da aba ao rodapé para ser impresso
    ws.oddFooter.center.text = "&A"

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtém a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 1) * 0.7  # Reduz em 50%
        ws.column_dimensions[column].width = adjusted_width
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=6)

    # Retorna o Workbook para ser salvo posteriormente
    return wb

# Função para criar o arquivo ZIP contendo os arquivos Excel
def criar_zip(arquivos_excel):
    buffer = BytesIO()
    with ZipFile(buffer, 'w') as zipf:
        for filename, wb in arquivos_excel.items():
            # Salva o workbook diretamente no arquivo ZIP
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            zipf.writestr(f"{filename}.xlsx", excel_buffer.read())
    buffer.seek(0)
    return buffer

# Interface do Streamlit
st.title("Processador de Listas CSV")

# Upload dos arquivos CSV
uploaded_files = st.file_uploader("Carregar arquivos CSV", accept_multiple_files=True, type="csv")

# Exibe os nomes dos arquivos carregados
if uploaded_files:
    st.write("Arquivos carregados:")
    for file in uploaded_files:
        st.write(file.name)

# Botão para processar os arquivos
if uploaded_files and st.button("Gerar Listas"):
    arquivos_excel = {}
    
    # Processa cada arquivo carregado
    for file in uploaded_files:
        df = pd.read_csv(file, sep=';')
        df_processado = processar_csv(df)
        wb = csv_para_excel(df_processado, file.name)
        arquivos_excel[file.name] = wb

    # Gera o arquivo ZIP
    zip_buffer = criar_zip(arquivos_excel)

    # Botão para baixar o ZIP
    st.download_button(
        label="Fazer Download",
        data=zip_buffer,
        file_name="listas_processadas.zip",
        mime="application/zip"
    )
